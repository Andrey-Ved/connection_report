from openpyxl import Workbook
from openpyxl.styles import (
    Alignment,
    Border,
    Font,
    numbers,
    PatternFill,
    Side,
)
from openpyxl.utils.cell import get_column_letter

from app.core.report_data import ReportDF
from app.core.settings import FilePath


def convert_to_xlsx_file(
        report: ReportDF,
        xls_report_file: FilePath,
) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "report"
    ws.freeze_panes = "B2"

    thins = Side(border_style="thin", color="000000")
    medium = Side(border_style="medium", color="000000")

    fill1 = PatternFill(fill_type='solid', fgColor="E2EFDA")  # noqa
    fill4 = PatternFill(fill_type='solid', fgColor="C6E0B4")  # noqa
    fill7 = PatternFill(fill_type='solid', fgColor="A9D08E")  # noqa
    fill10 = PatternFill(fill_type='solid', fgColor="548235") # noqa

    date_format = numbers.BUILTIN_FORMATS[14]

    columns_name = report.columns.tolist()

    for i in range(0, len(columns_name)):
        column_letter = get_column_letter(2 + i)
        ws.column_dimensions[column_letter].width = 12

        cell = ws.cell(1, 2 + i)

        cell.value = columns_name[i]
        cell.number_format = date_format
        cell.font = Font(name='Arial')
        cell.alignment = Alignment(horizontal='center')
        cell.border = Border(
            left=thins,
            right=thins,
            top=thins,
            bottom=medium,
        )

    ws.column_dimensions['A'].width = 25
    rows_name = report.index.tolist()

    for i in range(0, len(rows_name)):
        cell = ws.cell(2 + i, 1)

        cell.value = rows_name[i]
        cell.font = Font(name='Arial')
        cell.alignment = Alignment(horizontal='center')
        cell.border = Border(
            left=thins,
            right=medium,
            top=thins,
            bottom=thins,
        )

    for ir in range(0, len(report)):
        for ic in range(0, len(report.iloc[ir])):
            value = report.iloc[ir].iloc[ic]
            cell = ws.cell(2+ir, 2+ic)

            cell.value = value
            cell.font = Font(name='Arial')
            cell.alignment = Alignment(horizontal='center')
            cell.border = Border(
                left=thins,
                right=thins,
                top=thins,
                bottom=thins,
            )
            if value < 1:
                pass
            elif value < 4:
                cell.fill = fill1
            elif value < 7:
                cell.fill = fill4
            elif value < 10:
                cell.fill = fill7
            else:
                cell.fill = fill10

    wb.save(xls_report_file)
